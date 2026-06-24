from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
import os, html, psycopg2, sqlite3, json, uuid, time, datetime, base64, re, csv, io
from io import BytesIO
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "troque-esta-chave")

SENHA = os.environ.get("APP_SENHA", "escola1234")
DATABASE_URL = os.environ.get("DATABASE_URL")

# Configurações de cookies de sessão seguros
app.config.update(
    SESSION_COOKIE_SECURE=DATABASE_URL is not None,  # Apenas envia via HTTPS em produção
    SESSION_COOKIE_HTTPONLY=True,                    # Impede acesso JavaScript ao cookie
    SESSION_COOKIE_SAMESITE='Lax',                  # Protege contra CSRF
)

# Limitador de tentativas de login por IP (Anti-Brute Force)
LOGIN_LIMIT = 5
BLOCK_TIME = 300  # 5 minutos
failed_attempts = {}  # ip -> {"count": int, "blocked_until": float}

def check_rate_limit(ip):
    now = time.time()
    if ip in failed_attempts:
        record = failed_attempts[ip]
        if record["blocked_until"] > now:
            remaining = int(record["blocked_until"] - now)
            return False, f"Muitas tentativas falhas. Tente novamente em {remaining} segundos."
        elif record["count"] >= LOGIN_LIMIT:
            failed_attempts[ip] = {"count": 0, "blocked_until": 0.0}
    return True, ""

def register_login_failure(ip):
    now = time.time()
    if ip not in failed_attempts:
        failed_attempts[ip] = {"count": 0, "blocked_until": 0.0}
    failed_attempts[ip]["count"] += 1
    if failed_attempts[ip]["count"] >= LOGIN_LIMIT:
        failed_attempts[ip]["blocked_until"] = now + BLOCK_TIME

def reset_login_attempts(ip):
    if ip in failed_attempts:
        del failed_attempts[ip]

# Cabeçalhos de Segurança HTTP
@app.after_request
def add_security_headers(response):
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self' https://fonts.googleapis.com https://fonts.gstatic.com https://cdn.jsdelivr.net; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "  # CORRIGIDO: permite scripts inline
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net; "
        "font-src 'self' https://fonts.gstatic.com https://cdn.jsdelivr.net; "
        "img-src 'self' data:;"
    )
    return response

def get_conn():
    if DATABASE_URL:
        return psycopg2.connect(DATABASE_URL)
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "escola_local.db")
    return sqlite3.connect(db_path)

def bootstrap_admin():
    d = load()
    if "usuarios" not in d:
        d["usuarios"] = {}
    if not d["usuarios"]:
        admin_pass = os.environ.get("APP_SENHA", "escola1234")
        d["usuarios"]["admin"] = {
            "usuario": "admin",
            "nome": "Administrador",
            "senha_hash": generate_password_hash(admin_pass)
        }
        save(d)
        print("Usuário inicial 'admin' criado com sucesso.")

def init_db():
    conn = get_conn()
    c = conn.cursor()
    try:
        if DATABASE_URL:
            c.execute("""
                CREATE TABLE IF NOT EXISTS dados (
                    id INTEGER PRIMARY KEY DEFAULT 1,
                    conteudo JSONB NOT NULL
                )
            """)
            c.execute("""
                INSERT INTO dados (id, conteudo)
                VALUES (1, '{"alunos": {}, "materias": {}, "usuarios": {}, "turmas": {}}')
                ON CONFLICT (id) DO NOTHING
            """)
        else:
            c.execute("""
                CREATE TABLE IF NOT EXISTS dados (
                    id INTEGER PRIMARY KEY DEFAULT 1,
                    conteudo TEXT NOT NULL
                )
            """)
            c.execute("""
                INSERT OR IGNORE INTO dados (id, conteudo)
                VALUES (1, '{"alunos": {}, "materias": {}, "usuarios": {}, "turmas": {}}')
            """)
        conn.commit()
    finally:
        c.close()
        conn.close()
    bootstrap_admin()

def load():
    conn = get_conn()
    c = conn.cursor()
    try:
        c.execute("SELECT conteudo FROM dados WHERE id = 1")
        row = c.fetchone()
        if not row:
            return {"alunos": {}, "materias": {}, "usuarios": {}}
        val = row[0]
        if isinstance(val, str):
            d = json.loads(val)
        else:
            d = val
        if "usuarios" not in d:
            d["usuarios"] = {}
        if "turmas" not in d:
            d["turmas"] = {}
        if "config" not in d:
            d["config"] = {}
        if "professores" not in d:
            d["professores"] = {}
        if "cursos" not in d:
            d["cursos"] = {}
        return d
    finally:
        c.close()
        conn.close()

def save(dados):
    placeholder = "%s" if DATABASE_URL else "?"
    conn = get_conn()
    c = conn.cursor()
    try:
        c.execute(f"UPDATE dados SET conteudo = {placeholder} WHERE id = 1", [json.dumps(dados)])
        conn.commit()
    finally:
        c.close()
        conn.close()

def sanitize(text):
    return html.escape(str(text).strip())[:100]

def sanitize_long(text, max_len=2000):
    return html.escape(str(text).strip())[:max_len]

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logado"):
            return jsonify({"erro": "Não autorizado."}), 401
        return f(*args, **kwargs)
    return decorated

@app.route("/login", methods=["GET", "POST"])
def login():
    erro = ""
    if request.method == "POST":
        ip = request.remote_addr or "unknown"
        allowed, reason = check_rate_limit(ip)
        if not allowed:
            return render_template("login.html", erro=reason)

        usuario = sanitize(request.form.get("usuario", ""))
        senha = request.form.get("senha", "")

        if not usuario or not senha:
            erro = "Usuário e senha são obrigatórios."
            register_login_failure(ip)
        else:
            d = load()
            user_data = d.get("usuarios", {}).get(usuario)
            if user_data and check_password_hash(user_data["senha_hash"], senha):
                session.clear()
                session["logado"] = True
                session["usuario"] = usuario
                session["nome"] = user_data.get("nome", usuario)
                reset_login_attempts(ip)
                return redirect(url_for("index"))
            else:
                erro = "Usuário ou senha incorretos."
                register_login_failure(ip)

    return render_template("login.html", erro=erro)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
def index():
    if not session.get("logado"):
        return redirect(url_for("login"))
    return render_template("index.html")

@app.route("/api/dados")
@login_required
def get_dados():
    d = load()
    d.pop("usuarios", None)  # nunca expor hashes de senha ao frontend
    return jsonify(d)

@app.route("/api/alunos", methods=["POST"])
@login_required
def add_aluno():
    d = load()
    body = request.json
    nome = sanitize(body.get("nome", ""))
    turma = sanitize(body.get("turma", "")) if body.get("turma") else ""
    if not nome:
        return jsonify({"erro": "Nome é obrigatório."}), 400
    if any(a["nome"].lower() == nome.lower() for a in d["alunos"].values()):
        return jsonify({"erro": "Aluno já cadastrado."}), 400
    if turma and turma not in d["turmas"]:
        return jsonify({"erro": "Turma não encontrada."}), 404
    mat = str(uuid.uuid4())[:8]
    while mat in d["alunos"]:
        mat = str(uuid.uuid4())[:8]
    d["alunos"][mat] = {"nome": nome, "materias": [], "turma": turma}
    save(d)
    return jsonify({"ok": True})

@app.route("/api/alunos/importar", methods=["POST"])
@login_required
def importar_alunos():
    d = load()
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro": "Envie um arquivo .xlsx ou .csv."}), 400

    nome_arquivo = arquivo.filename.lower()
    turma_padrao = sanitize(request.form.get("turma", ""))
    if turma_padrao and turma_padrao not in d["turmas"]:
        return jsonify({"erro": "Turma padrão não encontrada."}), 404

    try:
        if nome_arquivo.endswith(".csv"):
            texto = arquivo.read().decode("utf-8-sig", errors="ignore")
            linhas_brutas = list(csv.reader(io.StringIO(texto)))
        elif nome_arquivo.endswith((".xlsx", ".xlsm")):
            wb = load_workbook(BytesIO(arquivo.read()), data_only=True)
            ws = wb.active
            linhas_brutas = [[c.value for c in row] for row in ws.iter_rows()]
        else:
            return jsonify({"erro": "Formato não suportado. Envie um arquivo .xlsx ou .csv."}), 400
    except Exception:
        return jsonify({"erro": "Não foi possível ler esse arquivo."}), 400

    linhas_brutas = [linha for linha in linhas_brutas if any(c not in (None, "") for c in linha)]
    if not linhas_brutas:
        return jsonify({"erro": "Arquivo vazio."}), 400

    primeira = [str(c).strip().lower() if c is not None else "" for c in linhas_brutas[0]]
    col_nome, col_turma, inicio = 0, 1, 0
    if "nome" in primeira:
        col_nome = primeira.index("nome")
        col_turma = primeira.index("turma") if "turma" in primeira else None
        inicio = 1

    nomes_existentes = {a["nome"].lower() for a in d["alunos"].values()}
    criados, duplicados, erros = 0, 0, []

    for i, linha in enumerate(linhas_brutas[inicio:], start=inicio + 1):
        if col_nome >= len(linha):
            continue
        nome_bruto = linha[col_nome]
        if nome_bruto is None or not str(nome_bruto).strip():
            continue
        nome = sanitize(str(nome_bruto))

        turma_linha = ""
        if col_turma is not None and col_turma < len(linha) and linha[col_turma]:
            turma_linha = sanitize(str(linha[col_turma]))
        turma_final = turma_linha or turma_padrao

        if turma_final and turma_final not in d["turmas"]:
            erros.append(f"Linha {i}: turma \"{turma_final}\" não existe — \"{nome}\" foi cadastrado sem turma.")
            turma_final = ""

        if nome.lower() in nomes_existentes:
            duplicados += 1
            continue

        mat = str(uuid.uuid4())[:8]
        while mat in d["alunos"]:
            mat = str(uuid.uuid4())[:8]
        d["alunos"][mat] = {"nome": nome, "materias": [], "turma": turma_final}
        nomes_existentes.add(nome.lower())
        criados += 1

    save(d)
    return jsonify({"ok": True, "criados": criados, "duplicados": duplicados, "erros": erros[:20]})

@app.route("/api/alunos/<matricula>", methods=["DELETE"])
@login_required
def del_aluno(matricula):
    d = load()
    matricula = sanitize(matricula)
    if matricula not in d["alunos"]:
        return jsonify({"erro": "Aluno não encontrado."}), 404
    del d["alunos"][matricula]
    for mat in d["materias"].values():
        mat.get("chamadas", {}).pop(matricula, None)
        mat.get("notas", {}).pop(matricula, None)
    save(d)
    return jsonify({"ok": True})

@app.route("/api/alunos/<matricula>", methods=["PUT"])
@login_required
def edit_aluno(matricula):
    d = load()
    matricula = sanitize(matricula)
    if matricula not in d["alunos"]:
        return jsonify({"erro": "Aluno não encontrado."}), 404
    body = request.json
    nome = sanitize(body.get("nome", ""))
    if not nome:
        return jsonify({"erro": "Nome é obrigatório."}), 400
    if any(a["nome"].lower() == nome.lower() and k != matricula for k, a in d["alunos"].items()):
        return jsonify({"erro": "Já existe um aluno com esse nome."}), 400
    d["alunos"][matricula]["nome"] = nome
    if "turma" in body:
        turma = sanitize(body.get("turma", "")) if body.get("turma") else ""
        if turma and turma not in d["turmas"]:
            return jsonify({"erro": "Turma não encontrada."}), 404
        d["alunos"][matricula]["turma"] = turma
    save(d)
    return jsonify({"ok": True})

@app.route("/api/alunos/<matricula>/materias", methods=["POST"])
@login_required
def assoc_materia(matricula):
    d = load()
    matricula = sanitize(matricula)
    materia = sanitize(request.json.get("materia", ""))
    if matricula not in d["alunos"]:
        return jsonify({"erro": "Aluno não encontrado."}), 404
    if materia not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    aluno = d["alunos"][matricula]
    if materia not in aluno["materias"]:
        aluno["materias"].append(materia)
    save(d)
    return jsonify({"ok": True})

@app.route("/api/alunos/<matricula>/materias/<materia>", methods=["DELETE"])
@login_required
def desassoc_materia(matricula, materia):
    d = load()
    aluno = d["alunos"].get(sanitize(matricula))
    materia = sanitize(materia)
    if aluno and materia in aluno["materias"]:
        aluno["materias"].remove(materia)
    save(d)
    return jsonify({"ok": True})

@app.route("/api/cursos", methods=["POST"])
@login_required
def add_curso():
    d = load()
    nome = sanitize(request.json.get("nome", ""))
    if not nome:
        return jsonify({"erro": "Nome do curso é obrigatório."}), 400
    if nome in d["cursos"]:
        return jsonify({"erro": "Curso já existe."}), 400
    d["cursos"][nome] = {}
    save(d)
    return jsonify({"ok": True})

@app.route("/api/cursos/<nome>", methods=["DELETE"])
@login_required
def del_curso(nome):
    d = load()
    nome = sanitize(nome)
    if nome not in d["cursos"]:
        return jsonify({"erro": "Curso não encontrado."}), 404
    del d["cursos"][nome]
    for mat in d["materias"].values():
        if mat.get("curso") == nome:
            mat["curso"] = ""
    save(d)
    return jsonify({"ok": True})

@app.route("/api/professores", methods=["POST"])
@login_required
def add_professor():
    d = load()
    nome = sanitize(request.json.get("nome", ""))
    if not nome:
        return jsonify({"erro": "Nome do professor é obrigatório."}), 400
    if nome in d["professores"]:
        return jsonify({"erro": "Professor já cadastrado."}), 400
    d["professores"][nome] = {}
    save(d)
    return jsonify({"ok": True})

@app.route("/api/professores/<nome>", methods=["DELETE"])
@login_required
def del_professor(nome):
    d = load()
    nome = sanitize(nome)
    if nome not in d["professores"]:
        return jsonify({"erro": "Professor não encontrado."}), 404
    del d["professores"][nome]
    for mat in d["materias"].values():
        if mat.get("professor") == nome:
            mat["professor"] = ""
    save(d)
    return jsonify({"ok": True})

@app.route("/api/professores/<nome>/materias", methods=["POST"])
@login_required
def atribuir_materias_professor(nome):
    d = load()
    nome = sanitize(nome)
    if nome not in d["professores"]:
        return jsonify({"erro": "Professor não encontrado."}), 404
    materias_novas = {sanitize(m) for m in request.json.get("materias", [])}
    for mat_nome, mat in d["materias"].items():
        if mat_nome in materias_novas:
            mat["professor"] = nome
        elif mat.get("professor") == nome:
            mat["professor"] = ""
    save(d)
    return jsonify({"ok": True})

@app.route("/api/materias/<nome>/curso", methods=["POST"])
@login_required
def definir_curso_materia(nome):
    d = load()
    nome = sanitize(nome)
    if nome not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    curso = sanitize(request.json.get("curso", "")) if request.json.get("curso") else ""
    if curso and curso not in d["cursos"]:
        return jsonify({"erro": "Curso não encontrado."}), 404
    d["materias"][nome]["curso"] = curso
    save(d)
    return jsonify({"ok": True})

@app.route("/api/materias/<nome>/professor", methods=["POST"])
@login_required
def definir_professor_materia(nome):
    d = load()
    nome = sanitize(nome)
    if nome not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    professor = sanitize(request.json.get("professor", "")) if request.json.get("professor") else ""
    if professor and professor not in d["professores"]:
        return jsonify({"erro": "Professor não encontrado."}), 404
    d["materias"][nome]["professor"] = professor
    save(d)
    return jsonify({"ok": True})

@app.route("/api/turmas/<turma>/associar-materia", methods=["POST"])
@login_required
def associar_turma_materia(turma):
    d = load()
    turma = sanitize(turma)
    materia = sanitize(request.json.get("materia", ""))
    if turma not in d["turmas"]:
        return jsonify({"erro": "Turma não encontrada."}), 404
    if materia not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    afetados = 0
    for aluno in d["alunos"].values():
        if aluno.get("turma") == turma and materia not in aluno.get("materias", []):
            aluno.setdefault("materias", []).append(materia)
            afetados += 1
    save(d)
    return jsonify({"ok": True, "afetados": afetados})

@app.route("/api/turmas/<turma>/desassociar-materia", methods=["POST"])
@login_required
def desassociar_turma_materia(turma):
    d = load()
    turma = sanitize(turma)
    materia = sanitize(request.json.get("materia", ""))
    if turma not in d["turmas"]:
        return jsonify({"erro": "Turma não encontrada."}), 404
    afetados = 0
    for aluno in d["alunos"].values():
        if aluno.get("turma") == turma and materia in aluno.get("materias", []):
            aluno["materias"].remove(materia)
            afetados += 1
    save(d)
    return jsonify({"ok": True, "afetados": afetados})

@app.route("/api/materias", methods=["POST"])
@login_required
def add_materia():
    d = load()
    nome = sanitize(request.json.get("nome", ""))
    prof = sanitize(request.json.get("professor", "")) if request.json.get("professor") else ""
    curso = sanitize(request.json.get("curso", "")) if request.json.get("curso") else ""
    if not nome:
        return jsonify({"erro": "Nome da matéria é obrigatório."}), 400
    if nome in d["materias"]:
        return jsonify({"erro": "Matéria já existe."}), 400
    if prof and prof not in d["professores"]:
        return jsonify({"erro": "Professor não encontrado."}), 404
    if curso and curso not in d["cursos"]:
        return jsonify({"erro": "Curso não encontrado."}), 404
    d["materias"][nome] = {"professor": prof, "curso": curso, "chamadas": {}, "notas": {}, "conteudos": {}}
    save(d)
    return jsonify({"ok": True})

@app.route("/api/materias/<nome>", methods=["DELETE"])
@login_required
def del_materia(nome):
    d = load()
    nome = sanitize(nome)
    if nome not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    del d["materias"][nome]
    for aluno in d["alunos"].values():
        if nome in aluno.get("materias", []):
            aluno["materias"].remove(nome)
    save(d)
    return jsonify({"ok": True})

@app.route("/api/materias/<nome>/conteudo", methods=["POST"])
@login_required
def add_conteudo(nome):
    d = load()
    nome = sanitize(nome)
    if nome not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    data = sanitize(request.json.get("data", ""))
    conteudo = sanitize_long(request.json.get("conteudo", ""))
    if not data:
        return jsonify({"erro": "Informe a data da aula."}), 400
    if not conteudo:
        return jsonify({"erro": "Descreva o conteúdo lecionado."}), 400
    d["materias"][nome].setdefault("conteudos", {})[data] = conteudo
    save(d)
    return jsonify({"ok": True})

@app.route("/api/materias/<nome>/conteudo/<data>", methods=["DELETE"])
@login_required
def del_conteudo(nome, data):
    d = load()
    nome = sanitize(nome)
    data = sanitize(data)
    if nome in d["materias"]:
        d["materias"][nome].setdefault("conteudos", {}).pop(data, None)
    save(d)
    return jsonify({"ok": True})

@app.route("/api/trocar-senha", methods=["POST"])
@login_required
def trocar_senha():
    d = load()
    usuario = session.get("usuario")
    user_data = d.get("usuarios", {}).get(usuario)
    if not user_data:
        return jsonify({"erro": "Usuário não encontrado."}), 404

    body = request.json or {}
    senha_atual = body.get("senha_atual", "")
    nova_senha = body.get("nova_senha", "")

    if not check_password_hash(user_data["senha_hash"], senha_atual):
        return jsonify({"erro": "Senha atual incorreta."}), 400
    if len(nova_senha) < 6:
        return jsonify({"erro": "A nova senha deve ter pelo menos 6 caracteres."}), 400

    d["usuarios"][usuario]["senha_hash"] = generate_password_hash(nova_senha)
    save(d)
    return jsonify({"ok": True})

@app.route("/api/logo", methods=["POST"])
@login_required
def upload_logo():
    d = load()
    imagem = (request.json or {}).get("imagem", "")
    m = re.match(r"^data:(image/(?:png|jpeg|jpg));base64,([A-Za-z0-9+/=]+)$", imagem)
    if not m:
        return jsonify({"erro": "Envie uma imagem PNG ou JPG válida."}), 400
    mime, b64data = m.group(1), m.group(2)
    if len(b64data) > 3_000_000:
        return jsonify({"erro": "Imagem muito grande. Use um arquivo de até 2MB."}), 400
    try:
        PILImage.open(BytesIO(base64.b64decode(b64data))).verify()
    except Exception:
        return jsonify({"erro": "Não foi possível ler essa imagem."}), 400
    d.setdefault("config", {})
    d["config"]["logo_mime"] = mime
    d["config"]["logo_b64"] = b64data
    save(d)
    return jsonify({"ok": True})

@app.route("/api/logo", methods=["DELETE"])
@login_required
def remover_logo():
    d = load()
    d.setdefault("config", {})
    d["config"].pop("logo_b64", None)
    d["config"].pop("logo_mime", None)
    save(d)
    return jsonify({"ok": True})

@app.route("/api/turmas", methods=["POST"])
@login_required
def add_turma():
    d = load()
    nome = sanitize(request.json.get("nome", ""))
    if not nome:
        return jsonify({"erro": "Nome da turma é obrigatório."}), 400
    if nome in d["turmas"]:
        return jsonify({"erro": "Turma já existe."}), 400
    d["turmas"][nome] = {}
    save(d)
    return jsonify({"ok": True})

@app.route("/api/turmas/<nome>", methods=["DELETE"])
@login_required
def del_turma(nome):
    d = load()
    nome = sanitize(nome)
    if nome not in d["turmas"]:
        return jsonify({"erro": "Turma não encontrada."}), 404
    del d["turmas"][nome]
    for aluno in d["alunos"].values():
        if aluno.get("turma") == nome:
            aluno["turma"] = ""
    save(d)
    return jsonify({"ok": True})

@app.route("/api/alunos/<matricula>/turma", methods=["POST"])
@login_required
def assoc_turma(matricula):
    d = load()
    matricula = sanitize(matricula)
    turma = sanitize(request.json.get("turma", ""))
    if matricula not in d["alunos"]:
        return jsonify({"erro": "Aluno não encontrado."}), 404
    if turma not in d["turmas"]:
        return jsonify({"erro": "Turma não encontrada."}), 404
    d["alunos"][matricula]["turma"] = turma
    save(d)
    return jsonify({"ok": True})

@app.route("/api/alunos/<matricula>/turma", methods=["DELETE"])
@login_required
def desassoc_turma(matricula):
    d = load()
    matricula = sanitize(matricula)
    if matricula not in d["alunos"]:
        return jsonify({"erro": "Aluno não encontrado."}), 404
    d["alunos"][matricula]["turma"] = ""
    save(d)
    return jsonify({"ok": True})

@app.route("/api/chamada", methods=["POST"])
@login_required
def salvar_chamada():
    d = load()
    body = request.json
    materia = sanitize(body.get("materia", ""))
    data = sanitize(body.get("data", ""))
    presencas = body.get("presencas", {})
    if materia not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    chamadas = d["materias"][materia].setdefault("chamadas", {})
    for mat, presente in presencas.items():
        chamadas.setdefault(sanitize(mat), {})[data] = bool(presente)
    save(d)
    return jsonify({"ok": True})

@app.route("/api/notas", methods=["POST"])
@login_required
def salvar_notas():
    d = load()
    body = request.json
    materia = sanitize(body.get("materia", ""))
    if materia not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    input_notas = body.get("notas", {})
    if not isinstance(input_notas, dict):
        return jsonify({"erro": "Dados inválidos."}), 400
    sanitized_notas = {}
    for mat_aluno, evals in input_notas.items():
        mat_aluno = sanitize(mat_aluno)
        if mat_aluno not in d["alunos"]:
            continue
        if not isinstance(evals, dict):
            continue
        aluno_grades = {}
        for aval, val in evals.items():
            aval = sanitize(aval)
            if not aval:
                continue
            if val is None or val == "":
                continue
            try:
                n = float(str(val).replace(",", "."))
                if 0 <= n <= 10:
                    aluno_grades[aval] = n
            except (ValueError, TypeError):
                pass
        if aluno_grades:
            sanitized_notas[mat_aluno] = aluno_grades
    d["materias"][materia]["notas"] = sanitized_notas
    save(d)
    return jsonify({"ok": True})

def _fmt_data(data_str, fmt="%d/%m"):
    try:
        return datetime.datetime.strptime(data_str, "%Y-%m-%d").strftime(fmt)
    except (ValueError, TypeError):
        return data_str

def gerar_diario_workbook(instituicao, curso, materia, professor, turma, ano, alunos, datas, chamadas, avals, notas, conteudos, logo_bytes=None):
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E2E8F0")

    wb = Workbook()
    ws = wb.active
    ws.title = "FRENTE"

    xl_logo = None
    col_offset = 0
    if logo_bytes:
        try:
            pil_img = PILImage.open(BytesIO(logo_bytes))
            largura, altura = pil_img.size
            altura_alvo = 70
            largura_alvo = max(1, int(largura * (altura_alvo / altura)))
            xl_logo = XLImage(BytesIO(logo_bytes))
            xl_logo.height = altura_alvo
            xl_logo.width = largura_alvo
            col_offset = 2
        except Exception:
            xl_logo = None
            col_offset = 0

    n_datas = len(datas)
    n_avals = len(avals)
    COL_NUM, COL_ALUNO = 1, 2
    col_first_data = 3
    col_first_aval = col_first_data + n_datas
    col_media = col_first_aval + n_avals
    col_faltas = col_media + 1
    total_cols = max(col_faltas, 5 + col_offset)

    title_start_col = 1 + col_offset
    ws.merge_cells(start_row=1, start_column=title_start_col, end_row=1, end_column=total_cols)
    c = ws.cell(1, title_start_col, "DIÁRIO DE CLASSE")
    c.font = title_font
    c.alignment = center

    rotulos = [(2, 1 + col_offset, "Instituição:", 2 + col_offset, instituicao or "—"), (2, 4 + col_offset, "Curso:", 5 + col_offset, curso or "—"),
               (3, 1 + col_offset, "Disciplina:", 2 + col_offset, materia), (3, 4 + col_offset, "Professor(a):", 5 + col_offset, professor or "—"),
               (4, 1 + col_offset, "Turma:", 2 + col_offset, turma), (4, 4 + col_offset, "Ano:", 5 + col_offset, str(ano))]
    for r, c1, label, c2, valor in rotulos:
        ws.cell(r, c1, label).font = bold
        ws.cell(r, c2, valor)

    label_secao = Alignment(horizontal="center", vertical="center", wrap_text=False)
    header_row = 6
    if n_datas:
        ws.merge_cells(start_row=header_row - 1, start_column=col_first_data, end_row=header_row - 1, end_column=col_first_data + n_datas - 1)
        c = ws.cell(header_row - 1, col_first_data, "FREQUÊNCIA")
        c.font = bold
        c.alignment = label_secao
        c.fill = header_fill
    if n_avals:
        ws.merge_cells(start_row=header_row - 1, start_column=col_first_aval, end_row=header_row - 1, end_column=col_first_aval + n_avals - 1)
        c = ws.cell(header_row - 1, col_first_aval, "AVALIAÇÕES")
        c.font = bold
        c.alignment = label_secao
        c.fill = header_fill

    ws.cell(header_row, COL_NUM, "Nº").font = bold
    ws.cell(header_row, COL_ALUNO, "Aluno").font = bold
    for i, data in enumerate(datas):
        cell = ws.cell(header_row, col_first_data + i, _fmt_data(data))
        cell.font = bold
        cell.alignment = center
    for i, av in enumerate(avals):
        cell = ws.cell(header_row, col_first_aval + i, av)
        cell.font = bold
        cell.alignment = center
    ws.cell(header_row, col_media, "Média").font = bold
    ws.cell(header_row, col_faltas, "Faltas").font = bold
    for col in range(1, col_faltas + 1):
        ws.cell(header_row, col).fill = header_fill
        ws.cell(header_row, col).border = border
        ws.cell(header_row, col).alignment = center

    row = header_row + 1
    for idx, (mat, a) in enumerate(alunos, start=1):
        ws.cell(row, COL_NUM, idx).alignment = center
        ws.cell(row, COL_ALUNO, a["nome"])
        chamadas_aluno = chamadas.get(mat, {})
        faltas = 0
        for i, data in enumerate(datas):
            presente = chamadas_aluno.get(data)
            if presente is True:
                val = "P"
            elif presente is False:
                val = "F"
                faltas += 1
            else:
                val = ""
            ws.cell(row, col_first_data + i, val).alignment = center
        notas_aluno = notas.get(mat, {})
        vals = []
        for i, av in enumerate(avals):
            v = notas_aluno.get(av)
            ws.cell(row, col_first_aval + i, v if v is not None else "").alignment = center
            if isinstance(v, (int, float)):
                vals.append(v)
        media = round(sum(vals) / len(vals), 2) if vals else ""
        ws.cell(row, col_media, media).alignment = center
        ws.cell(row, col_faltas, faltas).alignment = center
        for col in range(1, col_faltas + 1):
            ws.cell(row, col).border = border
        row += 1

    if n_datas:
        nota = ws.cell(row + 1, 1, "Legenda: P = presente   F = falta")
        nota.font = Font(italic=True, size=9, color="64748B")

    ws.column_dimensions[get_column_letter(COL_NUM)].width = 5
    ws.column_dimensions[get_column_letter(COL_ALUNO)].width = 28
    for i in range(n_datas):
        ws.column_dimensions[get_column_letter(col_first_data + i)].width = 5
    for i in range(n_avals):
        ws.column_dimensions[get_column_letter(col_first_aval + i)].width = 9
    ws.column_dimensions[get_column_letter(col_media)].width = 9
    ws.column_dimensions[get_column_letter(col_faltas)].width = 9
    ws.freeze_panes = ws.cell(header_row + 1, col_first_data)

    if xl_logo:
        ws.add_image(xl_logo, "A1")

    ws2 = wb.create_sheet("VERSO")
    ws2.merge_cells("A1:F1")
    c = ws2.cell(1, 1, "LANÇAMENTO DA MATÉRIA LECIONADA")
    c.font = title_font
    c.alignment = center

    ws2.cell(3, 1, "Data").font = bold
    ws2.cell(3, 2, "Conteúdo lecionado").font = bold
    ws2.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)
    for col in range(1, 7):
        ws2.cell(3, col).fill = header_fill
        ws2.cell(3, col).border = border

    r = 4
    for data in datas:
        ws2.cell(r, 1, _fmt_data(data, "%d/%m/%Y")).alignment = center
        ws2.cell(r, 2, conteudos.get(data, ""))
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        for col in range(1, 7):
            ws2.cell(r, col).border = border
        r += 1
    if not datas:
        r += 1

    r += 2
    ws2.cell(r, 1, "RESUMO DO BIMESTRE").font = bold
    r += 1
    ws2.cell(r, 1, "Aulas previstas:").font = bold
    r += 1
    ws2.cell(r, 1, "Aulas dadas:").font = bold
    ws2.cell(r, 3, len(datas))
    r += 1
    ws2.cell(r, 1, "Encerrado em:").font = bold
    r += 2
    ws2.cell(r, 1, "Professor(a):").font = bold
    ws2.cell(r, 3, professor or "")

    ws2.column_dimensions["A"].width = 18
    for col in "BCDEF":
        ws2.column_dimensions[col].width = 14

    return wb

@app.route("/api/diario", methods=["GET"])
@login_required
def gerar_diario():
    d = load()
    turma = sanitize(request.args.get("turma", ""))
    materia = sanitize(request.args.get("materia", ""))
    instituicao = sanitize(request.args.get("instituicao", ""))
    curso = sanitize(request.args.get("curso", ""))

    if materia not in d["materias"]:
        return jsonify({"erro": "Matéria não encontrada."}), 404
    if not turma or turma not in d["turmas"]:
        return jsonify({"erro": "Turma não encontrada."}), 404

    alunos = [(mat, a) for mat, a in d["alunos"].items()
              if materia in a.get("materias", []) and a.get("turma", "") == turma]
    alunos.sort(key=lambda x: x[1]["nome"])

    mat_dados = d["materias"][materia]
    if not curso:
        curso = mat_dados.get("curso", "")
    chamadas = mat_dados.get("chamadas", {})
    notas = mat_dados.get("notas", {})
    conteudos = mat_dados.get("conteudos", {})
    datas_chamada = {data for reg in chamadas.values() for data in reg.keys()}
    datas = sorted(datas_chamada | set(conteudos.keys()))
    avals = sorted({av for reg in notas.values() for av in reg.keys()})[:6]
    ano = datetime.datetime.now().year

    logo_bytes = None
    logo_b64 = d.get("config", {}).get("logo_b64")
    if logo_b64:
        try:
            logo_bytes = base64.b64decode(logo_b64)
        except Exception:
            logo_bytes = None

    wb = gerar_diario_workbook(instituicao, curso, materia, mat_dados.get("professor", ""), turma, ano, alunos, datas, chamadas, avals, notas, conteudos, logo_bytes)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_arquivo = f"Diario_{turma}_{materia}.xlsx".replace(" ", "_")
    return send_file(buf, as_attachment=True, download_name=nome_arquivo,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/backup", methods=["GET"])
@login_required
def exportar_backup():
    d = load()
    conteudo = json.dumps(d, ensure_ascii=False, indent=2)
    buf = BytesIO(conteudo.encode("utf-8"))
    nome_arquivo = f"backup_escola_{datetime.datetime.now().strftime('%Y-%m-%d_%H%M')}.json"
    return send_file(buf, as_attachment=True, download_name=nome_arquivo, mimetype="application/json")

@app.route("/api/backup/restaurar", methods=["POST"])
@login_required
def restaurar_backup():
    body = request.json or {}
    novo = body.get("dados")
    if not isinstance(novo, dict):
        return jsonify({"erro": "Arquivo inválido."}), 400

    chaves_esperadas = ("alunos", "materias", "turmas")
    if not all(isinstance(novo.get(k), dict) for k in chaves_esperadas):
        return jsonify({"erro": "Esse arquivo não parece ser um backup válido deste sistema."}), 400

    # Nunca restaura para um estado sem nenhum usuário (evita perder o acesso ao sistema)
    if not isinstance(novo.get("usuarios"), dict) or not novo.get("usuarios"):
        novo["usuarios"] = load().get("usuarios", {})
    if "config" not in novo or not isinstance(novo.get("config"), dict):
        novo["config"] = {}

    save(novo)
    return jsonify({"ok": True})

init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, port=port)
